import cv2
import os
import numpy as np
import pandas as pd
from tkinter import *
from tkinter import ttk, messagebox, simpledialog
from PIL import Image, ImageTk
import pickle
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class FaceRecognitionAttendance:
    def __init__(self, root):
        self.root = root
        self.root.title("Face Recognition Attendance System")
        self.root.geometry("800x500")

        self.camera_index = 0
        self.model_path = "TrainingImageLabel/Trainer.yml"
        self.label_path = "TrainingImageLabel/labels.pkl"
        self.attendance_file = "Attendance/attendance.xlsx"
        self.student_details_file = "StudentDetails/StudentDetails.xlsx"

        os.makedirs("TrainingImageLabel", exist_ok=True)
        os.makedirs("TrainingImage", exist_ok=True)
        os.makedirs("Attendance", exist_ok=True)
        os.makedirs("StudentDetails", exist_ok=True)

        Label(self.root, text="Face Recognition Attendance System", font=("Helvetica", 16, "bold"), fg="white", bg="blue").pack(side=TOP, fill=X)

        Button(self.root, text="Register & Capture Face", command=self.register_user, font=("Helvetica", 12, "bold")).pack(pady=10)
        Button(self.root, text="Train Model", command=self.train_model, font=("Helvetica", 12, "bold")).pack(pady=10)
        Button(self.root, text="Recognize & Mark Attendance", command=self.recognize_face, font=("Helvetica", 12, "bold")).pack(pady=10)
        Button(self.root, text="Switch Camera", command=self.switch_camera, font=("Helvetica", 12, "bold")).pack(pady=10)

        self.frame = Frame(self.root)
        self.frame.pack(pady=10)

        self.tree = ttk.Treeview(self.frame, columns=("UID", "Name", "Image Folder", "Status"), show="headings")
        for col in ("UID", "Name", "Image Folder", "Status"):
            self.tree.heading(col, text=col)
        self.tree.pack()

        self.tree.tag_configure("present", background="#C6EFCE")

        self.load_registered_users()
        self.refresh_treeview()

    def refresh_treeview(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        if os.path.exists(self.student_details_file):
            df = pd.read_excel(self.student_details_file, engine="openpyxl")
            for _, row in df.iterrows():
                tag = "present" if row["Status"] == "Present" else ""
                self.tree.insert("", "end", values=(row["UID"], row["Name"], row["Image_Folder"], row["Status"]), tags=(tag,))

    def load_registered_users(self):
        if os.path.exists(self.student_details_file):
            df = pd.read_excel(self.student_details_file, engine="openpyxl")
            if "Status" not in df.columns:
                df["Status"] = "Absent"
                df.to_excel(self.student_details_file, index=False, engine="openpyxl")

    def switch_camera(self):
        self.camera_index = 1 - self.camera_index
        messagebox.showinfo("Info", f"Switched to camera index: {self.camera_index}")

    def register_user(self):
        uid = simpledialog.askstring("Input", "Enter UID:")
        name = simpledialog.askstring("Input", "Enter Name:")
        if uid and name:
            self.capture_face(uid, name)
            self.save_to_excel(uid, name)
            self.refresh_treeview()

    def capture_face(self, uid, name):
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        cap = cv2.VideoCapture(self.camera_index)
        if not cap.isOpened():
            messagebox.showerror("Error", "Cannot open camera")
            return

        user_folder = os.path.join("TrainingImage", uid)
        os.makedirs(user_folder, exist_ok=True)

        count = 0
        while count < 10:
            ret, frame = cap.read()
            if not ret:
                messagebox.showerror("Error", "Failed to read frame from camera.")
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                face = gray[y:y+h, x:x+w]
                face = cv2.resize(face, (200, 200))
                cv2.imwrite(os.path.join(user_folder, f"{uid}_{count}.jpg"), face)
                count += 1

            cv2.imshow("Capturing Faces", frame)
            if cv2.waitKey(1) == 27:
                break

        cap.release()
        cv2.destroyAllWindows()
        messagebox.showinfo("Info", f"{count} face images captured for {name}")

    def save_to_excel(self, uid, name):
        df = pd.DataFrame([[uid, name, f"TrainingImage/{uid}", "Absent"]],
                          columns=["UID", "Name", "Image_Folder", "Status"])
        if os.path.exists(self.student_details_file):
            old_df = pd.read_excel(self.student_details_file, engine="openpyxl")
            df = pd.concat([old_df, df], ignore_index=True)
        df.to_excel(self.student_details_file, index=False, engine="openpyxl")

    def train_model(self):
        faces, labels = [], []
        label_map = {}
        label_id = 0

        for uid in os.listdir("TrainingImage"):
            folder_path = os.path.join("TrainingImage", uid)
            if not os.path.isdir(folder_path):
                continue

            if uid not in label_map:
                label_map[uid] = label_id
                label_id += 1

            for img_name in os.listdir(folder_path):
                img_path = os.path.join(folder_path, img_name)
                img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
                faces.append(img)
                labels.append(label_map[uid])

        if not faces:
            messagebox.showerror("Error", "No face data available for training.")
            return

        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.train(faces, np.array(labels))
        recognizer.save(self.model_path)

        with open(self.label_path, "wb") as f:
            pickle.dump(label_map, f)

        messagebox.showinfo("Info", "Model trained successfully!")

    def recognize_face(self):
        if not os.path.exists(self.model_path) or not os.path.exists(self.label_path):
            messagebox.showerror("Error", "Train the model first.")
            return

        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read(self.model_path)

        with open(self.label_path, "rb") as f:
            label_map = pickle.load(f)

        reverse_label_map = {v: k for k, v in label_map.items()}
        student_df = pd.read_excel(self.student_details_file, engine="openpyxl")

        cap = cv2.VideoCapture(self.camera_index)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

        attendance_data = []
        recognized_uid = None

        timeout = datetime.now().timestamp() + 5  # Run for 10 seconds

        while datetime.now().timestamp() < timeout:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                face = gray[y:y+h, x:x+w]
                face = cv2.resize(face, (200, 200))
                label, confidence = recognizer.predict(face)
                uid = reverse_label_map.get(label)

                if uid:
                    name_row = student_df[student_df["UID"] == uid]
                    if not name_row.empty:
                        name = name_row.iloc[0]["Name"]
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        attendance_data.append([uid, name, timestamp])
                        recognized_uid = uid

                        cv2.putText(frame, f"{name} ({uid})", (x, y - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

            cv2.imshow("Recognizing...", frame)
            if cv2.waitKey(1) == 5:
                break

        cap.release()
        cv2.destroyAllWindows()

        if attendance_data and recognized_uid:
            self.save_attendance(attendance_data)
            self.update_student_status(recognized_uid)

    def save_attendance(self, data):
        df = pd.DataFrame(data, columns=["UID", "Name", "DateTime"])
        if os.path.exists(self.attendance_file):
            old_df = pd.read_excel(self.attendance_file, engine="openpyxl")
            df = pd.concat([old_df, df], ignore_index=True)
        df.to_excel(self.attendance_file, index=False, engine="openpyxl")
        messagebox.showinfo("Info", "Attendance marked and saved successfully!")

    def update_student_status(self, uid):
        wb = load_workbook(self.student_details_file)
        ws = wb.active

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_col=4):  # Assuming UID in column A and Status in column D
            if str(row[0].value) == str(uid):
                row[3].value = "Present"
                for cell in row:
                    cell.fill = green_fill

        wb.save(self.student_details_file)
        self.refresh_treeview()


if __name__ == "__main__":
    root = Tk()
    app = FaceRecognitionAttendance(root)
    root.mainloop()
